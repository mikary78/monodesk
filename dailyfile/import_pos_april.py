"""
import_pos_april.py — 2026_04_pos.xlsx → sales_records POS 컬럼 업데이트
로컬 SQLite + Render PostgreSQL 동시 처리

POS 파일 컬럼 매핑 (row 6~35, 날짜 역순):
  col1: 날짜 (YYYY-MM-DD 형식)
  col4: 총거래금액 → pos_total
  col21: 순카드금액 (할인 등 차감 전) → pos_card
  col19: 현금금액 → pos_cash
  (col18 = 전체 합계, col20 = 현금영수증 별도)

실행: python dailyfile/import_pos_april.py
"""
import os
import sqlite3
import psycopg2
import openpyxl
from datetime import date, datetime

EXCEL_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH   = os.path.join(os.path.dirname(EXCEL_DIR), 'database', 'monodesk.db')
POS_PATH  = os.path.join(EXCEL_DIR, '2026_04_pos.xlsx')
DB_URL    = "postgresql://monodesk_user:oEC6WHpoWqT0RUWR4OOF242NayiWTufw@dpg-d7jdmqvlk1mc73a72krg-a.oregon-postgres.render.com/monodesk"

print(f'POS 파일: {POS_PATH}')
print(f'로컬 DB: {DB_PATH}')
print(f'Render DB: (연결 시도...)')
print()

def safe_float(v, default=0.0):
    if v is None: return default
    try: return float(v)
    except: return default

# POS 엑셀 파싱
wb = openpyxl.load_workbook(POS_PATH, data_only=True)
ws = wb.active

pos_data = {}  # { 'YYYY-MM-DD': {'pos_total': , 'pos_card': , 'pos_cash': } }

for row_num in range(6, ws.max_row + 1):
    raw_date = ws.cell(row_num, 1).value
    if raw_date is None:
        continue

    # 날짜 파싱
    if isinstance(raw_date, (date, datetime)):
        sales_date = raw_date.strftime('%Y-%m-%d')
    elif isinstance(raw_date, str) and len(raw_date) == 10:
        sales_date = raw_date  # 이미 YYYY-MM-DD 형식
    else:
        continue

    if not sales_date.startswith('2026-04-'):
        continue

    pos_total = safe_float(ws.cell(row_num, 4).value)
    pos_card  = safe_float(ws.cell(row_num, 21).value)  # 순카드
    pos_cash  = safe_float(ws.cell(row_num, 19).value)  # 현금

    pos_data[sales_date] = {
        'pos_total': pos_total,
        'pos_card':  pos_card,
        'pos_cash':  pos_cash,
    }

print(f'POS 파싱 완료: {len(pos_data)}일치 데이터')
for d in sorted(pos_data.keys()):
    v = pos_data[d]
    print(f'  {d}: 총 {int(v["pos_total"]):>8,}  카드 {int(v["pos_card"]):>8,}  현금 {int(v["pos_cash"]):>7,}')
print()

# ─── 로컬 SQLite UPDATE ─────────────────────────────────────────
print('=== 로컬 SQLite 업데이트 ===')
conn_sqlite = sqlite3.connect(DB_PATH)
cur_sqlite  = conn_sqlite.cursor()

updated_sqlite = 0
skipped_sqlite = 0

for sales_date, v in sorted(pos_data.items()):
    cur_sqlite.execute(
        'SELECT id FROM sales_records WHERE sales_date=? AND is_deleted=0',
        (sales_date,)
    )
    row = cur_sqlite.fetchone()
    if row:
        cur_sqlite.execute(
            """UPDATE sales_records
               SET pos_total=?, pos_card=?, pos_cash=?, updated_at=CURRENT_TIMESTAMP
               WHERE id=?""",
            (v['pos_total'], v['pos_card'], v['pos_cash'], row[0])
        )
        updated_sqlite += 1
        print(f'  {sales_date}: UPDATE OK (총 {int(v["pos_total"]):,})')
    else:
        skipped_sqlite += 1
        print(f'  {sales_date}: sales_records 없음 -> 스킵 (POS만 있음)')

conn_sqlite.commit()
conn_sqlite.close()
print(f'[OK] SQLite: {updated_sqlite}건 업데이트, {skipped_sqlite}건 스킵\n')

# ─── Render PostgreSQL UPDATE ────────────────────────────────────
print('=== Render PostgreSQL 업데이트 ===')
conn_pg = psycopg2.connect(DB_URL)
cur_pg  = conn_pg.cursor()

updated_pg = 0
skipped_pg = 0

for sales_date, v in sorted(pos_data.items()):
    cur_pg.execute(
        'SELECT id FROM sales_records WHERE sales_date=%s AND is_deleted=0',
        (sales_date,)
    )
    row = cur_pg.fetchone()
    if row:
        cur_pg.execute(
            """UPDATE sales_records
               SET pos_total=%s, pos_card=%s, pos_cash=%s, updated_at=NOW()
               WHERE id=%s""",
            (v['pos_total'], v['pos_card'], v['pos_cash'], row[0])
        )
        updated_pg += 1
        print(f'  {sales_date}: UPDATE OK (총 {int(v["pos_total"]):,})')
    else:
        skipped_pg += 1
        print(f'  {sales_date}: sales_records 없음 -> 스킵')

conn_pg.commit()
cur_pg.close()
conn_pg.close()
print(f'[OK] PostgreSQL: {updated_pg}건 업데이트, {skipped_pg}건 스킵\n')

print('=== POS import 완료 ===')
print(f'  SQLite     : {updated_sqlite}건')
print(f'  PostgreSQL : {updated_pg}건')
