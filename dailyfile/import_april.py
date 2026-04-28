"""
import_april.py — 여남동 4월 엑셀 데이터 → MonoDesk SQLite DB 삽입
실행: python dailyfile/import_april.py
"""
import os
import sqlite3
import openpyxl
from datetime import datetime

# ─── 경로 설정 ───────────────────────────────────────────────────────
EXCEL_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(os.path.dirname(EXCEL_DIR), 'database', 'monodesk.db')

xlsx_files = [f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx')]
if not xlsx_files:
    raise FileNotFoundError('dailyfile 폴더에 xlsx 파일이 없습니다.')
EXCEL_PATH = os.path.join(EXCEL_DIR, xlsx_files[0])
print(f'파일: {xlsx_files[0]}')
print(f'DB: {DB_PATH}\n')

# ─── 유틸리티 ────────────────────────────────────────────────────────
def safe_int(v, default=0):
    if v is None: return default
    try: return int(float(v))
    except: return default

def safe_float(v, default=0.0):
    if v is None: return default
    try: return float(v)
    except: return default

def safe_str(v):
    if v is None: return None
    s = str(v).strip()
    return s if s and s != '0' else None

def format_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    return None

# ─── 연결 ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

inserted = {'sales': 0, 'employees': 0, 'fixed_cost_items': 0, 'fixed_cost_records': 0}
skipped  = {'sales': 0, 'employees': 0, 'fixed_cost_items': 0}

# ═══════════════════════════════════════════════════════════════════════
# 1. 매출 시트 → sales_records
#
# 열 매핑 (0-indexed):
#   0: 일자   4: 카드순매출(catchtable 제외)  6: 현금매출
#   8: 할인   9: 서비스  10: 영수건수  11: 고객수
#  13: 배달매출  14: 이체건수  15: 이체금액
#  16: 캐치건수  17: 캐치금액  26: 취소건수  27: 취소액
#  28: 취소사유  29: 특이사항  31: 카드수수료  32: 배달수수료
# ═══════════════════════════════════════════════════════════════════════
print('─── 1. 매출 데이터 삽입 ───────────────────────')
ws = wb['1.매출']
YEAR, MONTH = 2026, 4

for row in ws.iter_rows(min_row=4, values_only=True):
    day = row[0]
    if not isinstance(day, int) or day <= 0:
        continue
    total = safe_int(row[1])
    if total == 0:
        print(f'  {day:02d}일: 총매출 0 → 스킵')
        continue

    sales_date = f'{YEAR}-{MONTH:02d}-{day:02d}'

    cur.execute(
        'SELECT id FROM sales_records WHERE sales_date=? AND is_deleted=0',
        (sales_date,)
    )
    if cur.fetchone():
        skipped['sales'] += 1
        print(f'  {day:02d}일: 이미 존재 → 스킵')
        continue

    card_net       = safe_float(row[4])   # 카드순매출 (catchtable 차감 완료)
    cash_col       = safe_float(row[6])   # 현금 (이체 포함)
    transfer_amt   = safe_float(row[15])  # 계좌이체 금액
    cash_pure      = cash_col - transfer_amt  # 순수 현금

    cur.execute('''
        INSERT INTO sales_records (
            sales_date,
            card_amount, cash_amount, cash_receipt_amount,
            delivery_amount, discount_amount, service_amount,
            receipt_count, customer_count,
            transfer_count, transfer_amount,
            catchtable_count, catchtable_amount,
            card_cancel_count, card_cancel_amount, card_cancel_reason,
            card_fee_estimated, delivery_fee_estimated,
            special_note,
            is_pos_synced, is_deleted, created_at, updated_at
        ) VALUES (
            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,
            0,0,CURRENT_TIMESTAMP,CURRENT_TIMESTAMP
        )
    ''', (
        sales_date,
        card_net,
        cash_pure,
        0,                       # cash_receipt (현금에 포함, 별도 미계상)
        safe_float(row[13]),     # delivery_amount
        safe_float(row[8]),      # discount_amount
        safe_float(row[9]),      # service_amount
        safe_int(row[10]),       # receipt_count
        safe_int(row[11]),       # customer_count
        safe_int(row[14]),       # transfer_count
        transfer_amt,
        safe_int(row[16]),       # catchtable_count
        safe_float(row[17]),     # catchtable_amount
        safe_int(row[26]),       # card_cancel_count
        safe_float(row[27]),     # card_cancel_amount
        safe_str(row[28]),       # card_cancel_reason
        safe_float(row[31]),     # card_fee_estimated
        safe_float(row[32]),     # delivery_fee_estimated
        safe_str(row[29]),       # special_note
    ))
    inserted['sales'] += 1
    print(f'  {day:02d}일: 총매출 {total:,}원 삽입 (카드 {int(card_net):,} / 현금 {int(cash_pure):,} / 캐치 {int(safe_float(row[17])):,})')

conn.commit()
print(f'[OK] 매출: {inserted["sales"]}건 삽입, {skipped["sales"]}건 스킵\n')


# ═══════════════════════════════════════════════════════════════════════
# 2. 직원현황 시트 → employees
#
# 주민번호, 계좌번호는 민감정보 — DB에 저장하지 않음
# ═══════════════════════════════════════════════════════════════════════
print('─── 2. 직원 데이터 삽입 ───────────────────────')
ws = wb['2-1.직원현황']

work_part_map = {'주방': 'kitchen', '홀': 'hall', '관리': 'management'}
employees_data = []
mode = None  # '4대보험' | '3.3%' | '파트타이머'

for row in ws.iter_rows(values_only=True):
    label = str(row[1]) if row[1] else ''

    # 섹션 헤더 감지
    if '4대보험' in label and '근로자' in label:
        mode = '4대보험'
        continue
    if '3.3%' in label:
        mode = '3.3%'
        continue
    if '파트타이머' in label and '현황' in label:
        mode = '파트타이머'
        continue

    if mode is None:
        continue

    # 이름 유효성 검사
    name = safe_str(row[2])
    if not name or name in ('구분', '이름', '선택'):
        continue

    if mode == '4대보험':
        position    = safe_str(row[4])          # 직책 (주방/홀)
        hire_date   = format_date(row[5])
        resign_date = format_date(row[6])
        work_cond   = safe_str(row[7])
        phone       = safe_str(row[9])           # 연락처 (OK)
        # col[10] = 주민번호 → 저장 안 함
        # col[11] = 계좌번호 → 저장 안 함
        monthly_sal = safe_float(row[13])
        meal_allow  = safe_int(row[16]) or 200000
        car_allow   = safe_int(row[17]) or 0

        if monthly_sal <= 0:
            continue  # 빈 템플릿 행

        wp = 'kitchen' if position == '주방' else 'hall' if position == '홀' else 'hall'
        employees_data.append({
            'name': name, 'employment_type': 'FULL_TIME',
            'salary_type': 'MONTHLY', 'contract_type': '4대보험',
            'has_insurance': 1, 'monthly_salary': monthly_sal, 'hourly_wage': None,
            'hire_date': hire_date, 'resign_date': resign_date,
            'position': position, 'work_part': wp, 'work_condition': work_cond,
            'phone': phone, 'meal_allowance': meal_allow, 'car_allowance': car_allow,
        })

    elif mode == '3.3%':
        hire_date   = format_date(row[5])
        phone       = safe_str(row[9])
        # col[10] = 주민번호 → 저장 안 함
        # col[11] = 계좌번호 → 저장 안 함
        monthly_sal = safe_float(row[13])

        if monthly_sal <= 0:
            continue

        employees_data.append({
            'name': name, 'employment_type': 'FULL_TIME',
            'salary_type': 'MONTHLY', 'contract_type': '3.3%',
            'has_insurance': 0, 'monthly_salary': monthly_sal, 'hourly_wage': None,
            'hire_date': hire_date, 'resign_date': None,
            'position': '매니저', 'work_part': 'management', 'work_condition': None,
            'phone': phone, 'meal_allowance': 200000, 'car_allowance': 0,
        })

    elif mode == '파트타이머':
        wp_str      = safe_str(row[3])
        phone       = safe_str(row[9])
        # col[10] = 주민번호 → 저장 안 함
        # col[11] = 계좌번호 → 저장 안 함
        hourly_wage = safe_float(row[8])

        if hourly_wage <= 0:
            continue

        wp = work_part_map.get(wp_str, 'hall')
        employees_data.append({
            'name': name, 'employment_type': 'PART_TIME',
            'salary_type': 'HOURLY', 'contract_type': '시급알바',
            'has_insurance': 0, 'monthly_salary': None, 'hourly_wage': hourly_wage,
            'hire_date': None, 'resign_date': None,
            'position': '파트타임', 'work_part': wp, 'work_condition': None,
            'phone': phone, 'meal_allowance': 0, 'car_allowance': 0,
        })

for emp in employees_data:
    cur.execute(
        'SELECT id FROM employees WHERE name=? AND is_deleted=0',
        (emp['name'],)
    )
    if cur.fetchone():
        skipped['employees'] += 1
        print(f'  {emp["name"]}: 이미 존재 → 스킵')
        continue

    cur.execute('''
        INSERT INTO employees (
            name, employment_type, salary_type, contract_type,
            has_insurance, monthly_salary, hourly_wage,
            hire_date, resign_date, position, work_part, work_condition,
            phone, meal_allowance, car_allowance,
            is_deleted, created_at, updated_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,CURRENT_TIMESTAMP,CURRENT_TIMESTAMP)
    ''', (
        emp['name'], emp['employment_type'], emp['salary_type'], emp['contract_type'],
        emp['has_insurance'], emp['monthly_salary'], emp['hourly_wage'],
        emp['hire_date'], emp['resign_date'], emp['position'],
        emp['work_part'], emp['work_condition'],
        emp['phone'], emp['meal_allowance'], emp['car_allowance'],
    ))
    inserted['employees'] += 1
    type_label = {'FULL_TIME': '정직원', 'PART_TIME': '파트타임'}[emp['employment_type']]
    print(f'  {emp["name"]} ({emp["contract_type"]}, {type_label}) 삽입')

conn.commit()
print(f'[OK] 직원: {inserted["employees"]}명 삽입, {skipped["employees"]}명 스킵\n')


# ═══════════════════════════════════════════════════════════════════════
# 3. 기초및고정비 시트 → fixed_cost_items + fixed_cost_records
# ═══════════════════════════════════════════════════════════════════════
print('─── 3. 고정비 데이터 삽입 ─────────────────────')
ws = wb['0.기초및고정비']

# (name, category, default_amount) 목록 수집
fixed_items = []
mode_f = None  # 'fixed1' | 'fixed2' | None

for row in ws.iter_rows(values_only=True):
    label = str(row[1]) if row[1] else ''

    if '고정비1' in label:
        mode_f = 'fixed1'
        continue
    if '고정비2' in label:
        mode_f = 'fixed2'
        continue
    if '총 합 계' in label:
        mode_f = None
        continue
    if '기초 세팅' in label or '매장 기본' in label:
        continue

    if mode_f is None:
        continue

    name = safe_str(row[1])
    if not name or name in ('비용구분',):
        continue

    default_amount = safe_int(row[4])  # 설정금액
    category = '임대료/주류대출' if mode_f == 'fixed1' else '운영고정비'

    fixed_items.append({
        'name': name,
        'category': category,
        'default_amount': default_amount,
    })

sort_order = 0
for item in fixed_items:
    cur.execute('SELECT id FROM fixed_cost_items WHERE name=?', (item['name'],))
    existing = cur.fetchone()

    if existing:
        item_id = existing[0]
        skipped['fixed_cost_items'] += 1
        print(f'  {item["name"]}: 이미 존재 → 스킵')
    else:
        cur.execute('''
            INSERT INTO fixed_cost_items (name, category, default_amount, is_active, sort_order, created_at)
            VALUES (?,?,?,1,?,CURRENT_TIMESTAMP)
        ''', (item['name'], item['category'], item['default_amount'], sort_order))
        item_id = cur.lastrowid
        inserted['fixed_cost_items'] += 1
        print(f'  {item["name"]}: {item["default_amount"]:,}원 삽입')
        sort_order += 1

    # 4월 고정비 기록 생성 (설정금액 > 0인 경우)
    if item['default_amount'] > 0:
        cur.execute(
            'SELECT id FROM fixed_cost_records WHERE item_id=? AND year=? AND month=?',
            (item_id, YEAR, MONTH)
        )
        if not cur.fetchone():
            cur.execute('''
                INSERT INTO fixed_cost_records
                    (item_id, year, month, default_amount, actual_amount, created_at, updated_at)
                VALUES (?,?,?,?,?,CURRENT_TIMESTAMP,CURRENT_TIMESTAMP)
            ''', (item_id, YEAR, MONTH, item['default_amount'], item['default_amount']))
            inserted['fixed_cost_records'] += 1
            print(f'    └ 4월 기록 생성 ({item["default_amount"]:,}원)')

conn.commit()
print(f'[OK] 고정비 항목: {inserted["fixed_cost_items"]}개 삽입, {skipped["fixed_cost_items"]}개 스킵')
print(f'[OK] 고정비 기록: {inserted["fixed_cost_records"]}건 삽입\n')

# ─── 최종 DB 현황 ─────────────────────────────────────────────────────
conn.close()
conn2 = sqlite3.connect(DB_PATH)
cur2 = conn2.cursor()

print('=' * 45)
print('삽입 완료 요약')
print('=' * 45)
print(f'  매출 기록       : {inserted["sales"]:>3}건 삽입 / {skipped["sales"]}건 스킵')
print(f'  직원             : {inserted["employees"]:>3}명 삽입 / {skipped["employees"]}명 스킵')
print(f'  고정비 항목      : {inserted["fixed_cost_items"]:>3}개 삽입 / {skipped["fixed_cost_items"]}개 스킵')
print(f'  고정비 기록      : {inserted["fixed_cost_records"]:>3}건 삽입')

print()
print('--- DB 전체 row 수 ---')
for table in ['sales_records', 'employees', 'fixed_cost_items', 'fixed_cost_records']:
    cur2.execute(f'SELECT COUNT(*) FROM {table}')
    print(f'  {table:<25}: {cur2.fetchone()[0]}')

# 매출 검증: 4월 매출 합계
cur2.execute('''
    SELECT COUNT(*), SUM(card_amount + cash_amount + catchtable_amount + transfer_amount)
    FROM sales_records
    WHERE sales_date LIKE '2026-04-%' AND is_deleted=0
''')
cnt, total_sum = cur2.fetchone()
print()
print('--- 4월 매출 검증 ---')
print(f'  레코드 수: {cnt}')
print(f'  매출 합계: {int(total_sum or 0):,}원')

conn2.close()
